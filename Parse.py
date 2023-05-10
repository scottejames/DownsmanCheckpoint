#!/usr/local/bin/python3

from ast import List
import datetime
import tkinter as tk
import openpyxl
import os
from Checkpoint import checkpoint

path = 'Templates'
save_path = 'Out'
sheetName = os.path.join(path, "LargeTemplate.xlsx")
wb = openpyxl.load_workbook(sheetName,data_only=True)
dataSheet = wb['Data']

# Find first cell in  column A that contains the word start
for i in range(1,40):
    cell = 'A' + str(i)
    if dataSheet[cell].value == 'Start':
        startRow = i
    if dataSheet[cell].value == 'Finish':
        finishRow = i

cpList = []

distance = dataSheet['D2'].value
hikeDate = dataSheet['D3'].value
mapList = dataSheet['D4'].value
for i in range(startRow, finishRow +1):
    number = dataSheet['A' + str(i)].value
    name = dataSheet['B' + str(i)].value    
    cp = checkpoint(name,number)
    cp.team = dataSheet['C' + str(i)].value 
    cp.mapReference = dataSheet['D' + str(i)].value
    cp.distanceFromPrior = dataSheet['E' + str(i)].value
    cp.timeAllowed = dataSheet['F' + str(i)].value
    cp.firstTeamArrival = dataSheet['G' + str(i)].value
    cp.firstTeamDeparture = dataSheet['H' + str(i)].value
    cp.lastTeamDeparture = dataSheet['K' + str(i)].value
    cpList.append(cp)


## Load some people

peopleSheet = wb['People']
lic = peopleSheet['B5'].value
licNumber = peopleSheet['C5'].value
districtNumber = peopleSheet['C6'].value
purpleCardName = peopleSheet['B7'].value
purpleCardNumber = peopleSheet['C7'].value
leadFirstAidName = peopleSheet['B8'].value
leadFirstAidNumber = peopleSheet['C8'].value
nap = peopleSheet['B9'].value
napNumber = peopleSheet['C9'].value
searchOneName = peopleSheet['B10'].value
searchOneNUmber = peopleSheet['C10'].value
searchtwoName = peopleSheet['B11'].value
searchTwoNumber = peopleSheet['C11'].value
searchThreeName = peopleSheet['B12'].value
searchThreeNumber = peopleSheet['C12'].value



#########################################
## Fill in the master sheet.
#########################################

cpSheet = wb["Master sheet"]
offset = 6
for i in range(0, len(cpList)):
    cp = cpList[i]
    if (i%2==0):
        cpSheet['A' + str(i + offset)].fill = openpyxl.styles.PatternFill(start_color='D0F9D9', end_color='D0F9D9', fill_type='solid')
        cpSheet['B' + str(i + offset)].fill = openpyxl.styles.PatternFill(start_color='D0F9D9', end_color='D0F9D9', fill_type='solid')
        cpSheet['C' + str(i + offset)].fill = openpyxl.styles.PatternFill(start_color='D0F9D9', end_color='D0F9D9', fill_type='solid')
        cpSheet['D' + str(i + offset)].fill = openpyxl.styles.PatternFill(start_color='D0F9D9', end_color='D0F9D9', fill_type='solid')
        cpSheet['E' + str(i + offset)].fill = openpyxl.styles.PatternFill(start_color='D0F9D9', end_color='D0F9D9', fill_type='solid')
        cpSheet['F' + str(i + offset)].fill = openpyxl.styles.PatternFill(start_color='D0F9D9', end_color='D0F9D9', fill_type='solid')
        cpSheet['G' + str(i + offset)].fill = openpyxl.styles.PatternFill(start_color='D0F9D9', end_color='D0F9D9', fill_type='solid')
        cpSheet['H' + str(i + offset)].fill = openpyxl.styles.PatternFill(start_color='D0F9D9', end_color='D0F9D9', fill_type='solid')
        cpSheet['I' + str(i + offset)].fill = openpyxl.styles.PatternFill(start_color='D0F9D9', end_color='D0F9D9', fill_type='solid')
        cpSheet['J' + str(i + offset)].fill = openpyxl.styles.PatternFill(start_color='D0F9D9', end_color='D0F9D9', fill_type='solid')
        cpSheet['K' + str(i + offset)].fill = openpyxl.styles.PatternFill(start_color='D0F9D9', end_color='D0F9D9', fill_type='solid')
    open = (datetime.datetime.combine(datetime.date.today(), cp.firstTeamArrival) - datetime.timedelta(minutes=30)).time()
    cpSheet['A' + str(i + offset)] = cp.number
    cpSheet['B' + str(i + offset)] = cp.name
    cpSheet['C' + str(i + offset)] = cp.mapReference
    cpSheet['D' + str(i + offset)] = 'Picture'
    cpSheet['E' + str(i + offset)] = cp.team
    cpSheet['F' + str(i + offset)] = 'Activity'
    cpSheet['G' + str(i + offset)] = open
    cpSheet['H' + str(i + offset)] = cp.distanceFromPrior
    cpSheet['I' + str(i + offset)] = cp.timeAllowed
    cpSheet['J' + str(i + offset)] = cp.firstTeamArrival
    cpSheet['K' + str(i + offset)] = cp.lastTeamDeparture
cpSheet['E1'] = hikeDate
cpSheet['D1'] = cpList[0].name

#########################################
## Fill in the Joining Instructions
#########################################

joinSheet = wb["Joining Instructions"]
startLocation = cpList[0].name
endLcoation = cpList[len(cpList)-1].name
endGridRef = cpList[len(cpList)-1].mapReference

parentsDropOffTime = (datetime.datetime.combine(datetime.date.today(), cpList[0].firstTeamArrival) - datetime.timedelta(minutes=30)).time()
earliestFinish = cpList[len(cpList)-1].firstTeamDeparture
latestFinish = cpList[len(cpList)-1].lastTeamDeparture
callHome = cpList[len(cpList)-1].timeAllowed

joinSheet['B5'] = startLocation
joinSheet['B6'] = endLcoation
joinSheet['B7'] = endGridRef
joinSheet['B9'] = parentsDropOffTime
joinSheet['B10'] = earliestFinish
joinSheet['B11'] = latestFinish
joinSheet['B12'] = callHome
joinSheet['B14'] = distance
joinSheet['B16'] = districtNumber
joinSheet['B18'] = mapList



## Save the completed workbook
resultsName = os.path.join(save_path, "Results.xlsx")
wb.save(resultsName)